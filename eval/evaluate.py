#!/usr/bin/env python3
"""Honest metrics for the A/B test results.

Reads the Excel produced by ab_test.py (ground truth + both systems'
verdicts) and prints accuracy, F1 per class, macro-F1, coverage (answered
vs abstained), a FEVER/LIAR breakdown, and a comparison against a naive
"always true" baseline. It does NOT touch the system — read-only analysis.

Rules:
  - MANIPULATION / MIXED count as "false";
  - UNVERIFIABLE (and ERROR/empty) = abstention: excluded from accuracy/F1,
    reported separately as coverage (how often the system stayed silent).

Usage:
    python3 eval/evaluate.py                 # reads ab_results.xlsx
    python3 eval/evaluate.py path/to.xlsx
"""

import sys
from collections import defaultdict
from typing import List, Optional


def to_binary(verdict: str) -> Optional[str]:
    """Map a verdict to 'true' / 'false' / None (abstention)."""
    v = (verdict or "").strip().lower()
    if v == "true":
        return "true"
    if v in ("false", "manipulation", "mixed"):
        return "false"
    return None  # unverifiable / error / empty -> abstain


def prf(preds: List[Optional[str]], golds: List[str], positive: str):
    """Precision/recall/F1 for one class, over answered items only."""
    tp = fp = fn = 0
    for p, g in zip(preds, golds):
        if p is None:
            continue
        if p == positive and g == positive:
            tp += 1
        elif p == positive and g != positive:
            fp += 1
        elif p != positive and g == positive:
            fn += 1
    prec = tp / (tp + fp) if (tp + fp) else 0.0
    rec = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0
    return prec, rec, f1


def report(name: str, preds: List[Optional[str]], golds: List[str]) -> None:
    total = len(golds)
    answered = [(p, g) for p, g in zip(preds, golds) if p is not None]
    n_ans = len(answered)
    correct = sum(1 for p, g in answered if p == g)

    print(f"\n=== {name} ===")
    print(f"  Всего: {total} | ответил: {n_ans} | смолчал (UNVERIFIABLE): {total - n_ans}")
    print(f"  Coverage (доля ответов): {100 * n_ans / total:.1f}%")
    if n_ans:
        print(f"  Точность среди ответов: {100 * correct / n_ans:.1f}%  ({correct}/{n_ans})")
    print(f"  Точность по всей выборке (молчание = ошибка): "
          f"{100 * correct / total:.1f}%  ({correct}/{total})")

    pt, rt, f1t = prf(preds, golds, "true")
    pf, rf, f1f = prf(preds, golds, "false")
    macro = (f1t + f1f) / 2
    print(f"  F1(true)={f1t:.2f}  F1(false)={f1f:.2f}  macro-F1={macro:.2f}")


def report_by_dataset(name: str, rows) -> None:
    groups = defaultdict(lambda: ([], []))
    for ds, pred, gold in rows:
        groups[ds or "unknown"][0].append(pred)
        groups[ds or "unknown"][1].append(gold)
    for ds in sorted(groups):
        preds, golds = groups[ds]
        answered = [(p, g) for p, g in zip(preds, golds) if p is not None]
        correct = sum(1 for p, g in answered if p == g)
        cov = 100 * len(answered) / len(golds) if golds else 0
        acc = 100 * correct / len(answered) if answered else 0
        print(f"    {name} / {ds}: coverage {cov:.0f}%, "
              f"точность среди ответов {acc:.0f}% ({correct}/{len(answered)})")


def main() -> None:
    path = sys.argv[1] if len(sys.argv) > 1 else "ab_results.xlsx"
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Нужна библиотека openpyxl:  pip3 install openpyxl")
        sys.exit(1)

    try:
        ws = load_workbook(path)["results"]
    except FileNotFoundError:
        print(f"Файл не найден: {path}\nСначала запусти ab_test.py.")
        sys.exit(1)

    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {name: i for i, name in enumerate(header)}
    needed = ["ground_truth", "baseline_verdict", "detector_verdict", "dataset"]
    for col in needed:
        if col not in idx:
            print(f"В файле нет колонки '{col}'. Это результат ab_test.py?")
            sys.exit(1)

    golds, base_preds, det_preds = [], [], []
    base_rows, det_rows = [], []
    for r in rows[1:]:
        gt = (r[idx["ground_truth"]] or "").strip().lower()
        if gt not in ("true", "false"):
            continue  # no usable ground truth
        ds = r[idx["dataset"]]
        b = to_binary(r[idx["baseline_verdict"]])
        d = to_binary(r[idx["detector_verdict"]])
        golds.append(gt)
        base_preds.append(b)
        det_preds.append(d)
        base_rows.append((ds, b, gt))
        det_rows.append((ds, d, gt))

    n = len(golds)
    if not n:
        print("Нет утверждений с эталонным вердиктом true/false.")
        sys.exit(1)

    print("=" * 60)
    print(f"ОЦЕНКА КАЧЕСТВА  (файл: {path}, утверждений с эталоном: {n})")
    print("Правила: MANIPULATION/MIXED = ложь; UNVERIFIABLE = молчание (вне точности/F1)")
    print("=" * 60)

    # Naive baseline: always predict the majority class.
    majority = "true" if golds.count("true") >= golds.count("false") else "false"
    naive_correct = golds.count(majority)
    print(f"\n=== Наивный baseline ('всегда {majority}') ===")
    print(f"  Точность: {100 * naive_correct / n:.1f}%  ({naive_correct}/{n})")

    report("Обычная LLM (baseline)", base_preds, golds)
    report_by_dataset("baseline", base_rows)
    report("Агент-детектор", det_preds, golds)
    report_by_dataset("detector", det_rows)


if __name__ == "__main__":
    main()
