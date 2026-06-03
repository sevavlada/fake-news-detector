#!/usr/bin/env python3
"""Batch A/B test runner.

Reads a JSON file of claims, runs EACH claim through both systems:
  - Baseline   : the single-LLM fact-checker (baseline_llm/)
  - Detector   : the multi-agent Fake News Detector, Architecture B
                 (all agents in sequence: D -> T -> C -> synthesizer)
and writes the side-by-side results to an Excel .xlsx file so they can be
compared in an A/B test. If the input contains a ground-truth `verdict`
(e.g. FEVER/LIAR datasets), accuracy of each system is also computed.

Usage:
    python3 ab_test.py claims.json
    python3 ab_test.py claims.json -o results.xlsx
    python3 ab_test.py claims.json --limit 5            # quick check
    python3 ab_test.py claims.json --field claim        # send plain claim
    python3 ab_test.py claims.json --field claim_with_context  # default

Accepted JSON shapes:
    ["claim 1", "claim 2", ...]
    [{"claim": "..."}, {"statement": "..."}, ...]
    {"claims": [ ... ]}   (or any object whose value is the list)
Dataset objects may also carry "verdict" (ground truth), "uid", "dataset".
"""

import argparse
import json
import os
import sys
from typing import Any, Dict, List

# Make the package importable (mirrors run.py).
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PARENT_DIR = os.path.dirname(_THIS_DIR)
if _PARENT_DIR not in sys.path:
    sys.path.insert(0, _PARENT_DIR)

from fake_news_detector.state import create_initial_state
from fake_news_detector.graphs.architecture_b import create_parallel_graph
from fake_news_detector.utils import state_to_verdict, format_protocol
from fake_news_detector.baseline_llm.chat import check_claim as baseline_check

# Keys we look for when claims are objects rather than plain strings.
_CLAIM_KEYS = ("claim", "statement", "text", "sentence", "utterance", "content")


def _claim_text(item: Any, field: str) -> str:
    """Pick the text to fact-check from one JSON element."""
    if isinstance(item, str):
        return item.strip()
    if isinstance(item, dict):
        # Preferred field first (e.g. claim_with_context), then the usual keys.
        for key in (field, *_CLAIM_KEYS):
            if key in item and str(item[key]).strip():
                return str(item[key]).strip()
        for value in item.values():
            if isinstance(value, str) and value.strip():
                return value.strip()
    return str(item).strip()


def load_records(path: str, field: str) -> List[Dict[str, Any]]:
    """Load claims as records: {claim, ground_truth, uid, dataset}."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    # Unwrap {"claims": [...]} / {"data": [...]} -> the inner list.
    if isinstance(data, dict):
        for value in data.values():
            if isinstance(value, list):
                data = value
                break

    if not isinstance(data, list):
        raise ValueError(
            "JSON must be a list of claims (or an object containing one). "
            f"Got: {type(data).__name__}"
        )

    records = []
    for idx, item in enumerate(data, start=1):
        text = _claim_text(item, field)
        if not text:
            continue
        meta = item if isinstance(item, dict) else {}
        records.append({
            "claim": text,
            "ground_truth": str(meta.get("verdict", "")).strip().lower(),
            "uid": meta.get("uid", idx),
            "dataset": meta.get("dataset", ""),
        })
    return records


def run_detector(claim: str) -> Dict[str, Any]:
    """Run one claim through the multi-agent detector (Architecture B).

    Returns the full graph state so we can surface the whole reasoning chain.
    """
    graph = create_parallel_graph()
    return graph.invoke(create_initial_state(claim))


def _factors(verdict: Dict[str, Any]) -> str:
    return "; ".join(verdict.get("key_factors", []))


def _correct(verdict: str, ground_truth: str) -> str:
    """Return YES/NO/'' comparing a system verdict to ground truth."""
    if not ground_truth:
        return ""
    return "YES" if str(verdict).strip().lower() == ground_truth else "NO"


def _agent_fields(state: Dict[str, Any]) -> Dict[str, Any]:
    """Pull each agent's report out of the state for the results table."""
    proto = (state or {}).get("protocol", {}) or {}

    def _d(key):
        v = proto.get(key)
        return v if isinstance(v, dict) else {}

    d, t, c, syn = _d("D"), _d("T"), _d("C"), _d("synthesis")
    d_sources = "; ".join(s.get("url", "") for s in d.get("sources", []) if s.get("url"))
    return {
        "agentD_verdict": d.get("verdict", ""),
        "agentD_confidence": d.get("confidence", ""),
        "agentD_sources": d_sources,
        "agentD_reasoning": d.get("reasoning", ""),
        "agentT_manipulation": t.get("manipulation_score", ""),
        "agentT_flags": "; ".join(t.get("flags", []) or []),
        "agentT_reasoning": t.get("reasoning", ""),
        "agentC_risk": c.get("risk_level", ""),
        "agentC_reasoning": c.get("reasoning", ""),
        "synthesis_reasoning": syn.get("reasoning", ""),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Batch A/B test: baseline vs. detector.")
    parser.add_argument("input", help="Path to the JSON file with claims")
    parser.add_argument("-o", "--output", default="ab_results.xlsx",
                        help="Output Excel file (default: ab_results.xlsx)")
    parser.add_argument("--limit", type=int, default=0,
                        help="Only process the first N claims (0 = all)")
    parser.add_argument("--field", default="claim_with_context",
                        help="Which field to send to the models "
                             "(default: claim_with_context; use 'claim' for the bare statement)")
    args = parser.parse_args()

    try:
        from openpyxl import Workbook
    except ImportError:
        print("Missing dependency. Install it with:\n    pip3 install openpyxl")
        sys.exit(1)

    records = load_records(args.input, args.field)
    if args.limit > 0:
        records = records[: args.limit]
    if not records:
        print("No claims found in the input file.")
        sys.exit(1)

    print(f"Loaded {len(records)} claim(s). Field: '{args.field}'. Running A/B test...\n")

    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    headers = [
        "uid", "dataset", "claim", "ground_truth",
        "baseline_verdict", "baseline_confidence", "baseline_correct", "baseline_reasoning",
        "detector_verdict", "detector_confidence", "detector_correct", "verdicts_match",
        # --- full detector reasoning chain (XAI), same as a single run ---
        "agentD_verdict", "agentD_confidence", "agentD_sources", "agentD_reasoning",
        "agentT_manipulation", "agentT_flags", "agentT_reasoning",
        "agentC_risk", "agentC_reasoning",
        "synthesis_reasoning",
        "detector_full_protocol",
    ]
    ws.append(headers)

    base_correct = det_correct = scored = match = 0

    for i, rec in enumerate(records, start=1):
        claim, gt = rec["claim"], rec["ground_truth"]
        print(f"[{i}/{len(records)}] {claim[:70]}")

        try:
            base = baseline_check(claim)
        except Exception as e:  # noqa: BLE001
            base = {"verdict": "ERROR", "confidence": 0, "key_factors": [],
                    "reasoning": f"baseline error: {e}"}
        try:
            det_state = run_detector(claim)
            det = state_to_verdict(det_state)
            fields = _agent_fields(det_state)
            protocol_text = format_protocol(det_state)
        except Exception as e:  # noqa: BLE001
            det = {"verdict": "ERROR", "confidence": 0, "reasoning": f"detector error: {e}"}
            fields = _agent_fields(None)
            protocol_text = f"detector error: {e}"

        bc = _correct(base.get("verdict"), gt)
        dc = _correct(det.get("verdict"), gt)
        same = base.get("verdict") == det.get("verdict")

        if gt:
            scored += 1
            base_correct += (bc == "YES")
            det_correct += (dc == "YES")
        match += same

        ws.append([
            rec["uid"], rec["dataset"], claim, gt,
            base.get("verdict"), base.get("confidence"), bc, base.get("reasoning"),
            det.get("verdict"), det.get("confidence"), dc, "YES" if same else "NO",
            fields["agentD_verdict"], fields["agentD_confidence"],
            fields["agentD_sources"], fields["agentD_reasoning"],
            fields["agentT_manipulation"], fields["agentT_flags"], fields["agentT_reasoning"],
            fields["agentC_risk"], fields["agentC_reasoning"],
            fields["synthesis_reasoning"],
            protocol_text,
        ])

    # Readability: wrap long text columns, widen them, freeze the header row.
    wide = {"claim": 45, "baseline_reasoning": 55, "agentD_sources": 45,
            "agentD_reasoning": 55, "agentT_reasoning": 45, "agentC_reasoning": 45,
            "synthesis_reasoning": 60, "detector_full_protocol": 90}
    for idx, name in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = wide.get(name, 16)
        if name in wide:
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # Summary sheet.
    total = len(records)
    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])
    summary.append(["total claims", total])
    summary.append(["scored against ground truth", scored])
    if scored:
        summary.append(["baseline correct", base_correct])
        summary.append(["baseline accuracy", round(100 * base_correct / scored, 1)])
        summary.append(["detector correct", det_correct])
        summary.append(["detector accuracy", round(100 * det_correct / scored, 1)])
    summary.append(["baseline/detector verdicts match", match])
    summary.append(["agreement rate %", round(100 * match / total, 1)])

    wb.save(args.output)

    print(f"\nDone. Results written to: {args.output}")
    if scored:
        print(f"  Baseline accuracy: {100 * base_correct / scored:.1f}%  "
              f"({base_correct}/{scored})")
        print(f"  Detector accuracy: {100 * det_correct / scored:.1f}%  "
              f"({det_correct}/{scored})")
    print(f"  Verdicts match:    {100 * match / total:.1f}%  ({match}/{total})")


if __name__ == "__main__":
    main()
